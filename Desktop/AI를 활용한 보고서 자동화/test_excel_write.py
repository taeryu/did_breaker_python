#!/usr/bin/env python3
# Excel 쓰기 테스트

from openpyxl import load_workbook
import shutil

def test_excel_write():
    """Excel 파일 쓰기 테스트"""
    
    print("📝 Excel 쓰기 테스트 시작...")
    
    try:
        # 템플릿 복사
        source = "재무상태표템플릿.xlsx"
        test_file = "재무상태표_테스트결과.xlsx"
        shutil.copy2(source, test_file)
        
        # 워크북 열기
        wb = load_workbook(test_file)
        ws = wb.active  # 첫 번째 시트
        
        print(f"시트명: {ws.title}")
        
        # 테스트 데이터 입력
        test_data = {
            'B5': 50000000,    # 현금및현금성자산
            'B6': 30000000,    # 단기금융상품  
            'B7': 100000000,   # 매출채권
            'D5': 40000000,    # 매입채무
            'D6': 25000000,    # 기타채무
            'D7': 60000000     # 단기차입금
        }
        
        print("데이터 입력 중...")
        for cell_addr, value in test_data.items():
            ws[cell_addr] = value
            print(f"  {cell_addr} = {value:,}")
        
        # 저장
        wb.save(test_file)
        print(f"✅ 파일 저장 완료: {test_file}")
        
        # 결과 확인
        wb2 = load_workbook(test_file)
        ws2 = wb2.active
        
        print("\n📊 입력 결과 확인:")
        for cell_addr in test_data.keys():
            value = ws2[cell_addr].value
            print(f"  {cell_addr}: {value}")
            
        print("\n🎉 Excel 쓰기 테스트 성공!")
        
    except Exception as e:
        print(f"💥 테스트 실패: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_write()