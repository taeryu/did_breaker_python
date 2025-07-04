# 셀매핑.xlsx 샘플 파일 생성 코드
# 이 코드를 실행하면 "셀매핑.xlsx" 파일이 생성됩니다.

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_cell_mapping_excel():
    """셀매핑 Excel 파일 생성"""
    
    # 매핑 데이터 정의
    mapping_data = [
        # [파일명, 섹션, 계정명, 셀주소, 데이터소스, 제목셀, 제목템플릿]
        ["경영실적요약", "당월데이터", "매출액", "B5", "IS", "A1", "{year}년 {month}월 경영실적요약"],
        ["경영실적요약", "당월데이터", "매출원가", "B6", "IS", "", ""],
        ["경영실적요약", "당월데이터", "매출총이익", "B7", "IS", "", ""],
        ["경영실적요약", "당월데이터", "판매비와관리비", "B8", "IS", "", ""],
        ["경영실적요약", "당월데이터", "영업이익", "B9", "IS", "", ""],
        ["경영실적요약", "당월데이터", "영업외수익", "B10", "IS", "", ""],
        ["경영실적요약", "당월데이터", "영업외비용", "B11", "IS", "", ""],
        ["경영실적요약", "당월데이터", "법인세비용차감전순이익", "B12", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "매출액", "C5", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "매출원가", "C6", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "매출총이익", "C7", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "판매비와관리비", "C8", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "영업이익", "C9", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "영업외수익", "C10", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "영업외비용", "C11", "IS", "", ""],
        ["경영실적요약", "전년동월데이터", "법인세비용차감전순이익", "C12", "IS", "", ""],
        
        ["반반장표", "판관비상세", "급여", "B10", "IS", "", ""],
        ["반반장표", "판관비상세", "상여금", "B11", "IS", "", ""],
        ["반반장표", "판관비상세", "퇴직급여", "B12", "IS", "", ""],
        ["반반장표", "판관비상세", "복리후생비", "B13", "IS", "", ""],
        ["반반장표", "판관비상세", "여비교통비", "B14", "IS", "", ""],
        ["반반장표", "판관비상세", "통신비", "B15", "IS", "", ""],
        ["반반장표", "판관비상세", "세금과공과", "B16", "IS", "", ""],
        ["반반장표", "판관비상세", "감가상각비", "B17", "IS", "", ""],
        ["반반장표", "판관비상세", "지급임차료", "B18", "IS", "", ""],
        ["반반장표", "판관비상세", "보험료", "B19", "IS", "", ""],
        ["반반장표", "판관비상세", "접대비", "B20", "IS", "", ""],
        ["반반장표", "판관비상세", "광고선전비", "B21", "IS", "", ""],
        ["반반장표", "판관비상세", "차량유지비", "B22", "IS", "", ""],
        ["반반장표", "판관비상세", "소모품비", "B23", "IS", "", ""],
        ["반반장표", "판관비상세", "용역비", "B24", "IS", "", ""],
        ["반반장표", "판관비상세", "기타판관비", "B25", "IS", "", ""],
        
        ["반반장표", "영업외손익상세", "이자수익", "B30", "IS", "", ""],
        ["반반장표", "영업외손익상세", "배당금수익", "B31", "IS", "", ""],
        ["반반장표", "영업외손익상세", "외환차익", "B32", "IS", "", ""],
        ["반반장표", "영업외손익상세", "기타영업외수익", "B33", "IS", "", ""],
        ["반반장표", "영업외손익상세", "이자비용", "B35", "IS", "", ""],
        ["반반장표", "영업외손익상세", "외환차손", "B36", "IS", "", ""],
        ["반반장표", "영업외손익상세", "기타영업외비용", "B37", "IS", "", ""],
        
        ["재무상태표", "유동자산", "현금및현금성자산", "B5", "BS", "A1", "{year}년 {month}월말 재무상태표"],
        ["재무상태표", "유동자산", "단기금융상품", "B6", "BS", "", ""],
        ["재무상태표", "유동자산", "매출채권", "B7", "BS", "", ""],
        ["재무상태표", "유동자산", "기타채권", "B8", "BS", "", ""],
        ["재무상태표", "유동자산", "재고자산", "B9", "BS", "", ""],
        ["재무상태표", "유동자산", "선급금", "B10", "BS", "", ""],
        ["재무상태표", "유동자산", "선급비용", "B11", "BS", "", ""],
        
        ["재무상태표", "비유동자산", "장기금융상품", "B15", "BS", "", ""],
        ["재무상태표", "비유동자산", "투자자산", "B16", "BS", "", ""],
        ["재무상태표", "비유동자산", "유형자산", "B17", "BS", "", ""],
        ["재무상태표", "비유동자산", "무형자산", "B18", "BS", "", ""],
        ["재무상태표", "비유동자산", "기타비유동자산", "B19", "BS", "", ""],
        
        ["재무상태표", "유동부채", "매입채무", "D5", "BS", "", ""],
        ["재무상태표", "유동부채", "기타채무", "D6", "BS", "", ""],
        ["재무상태표", "유동부채", "단기차입금", "D7", "BS", "", ""],
        ["재무상태표", "유동부채", "미지급금", "D8", "BS", "", ""],
        ["재무상태표", "유동부채", "미지급비용", "D9", "BS", "", ""],
        ["재무상태표", "유동부채", "선수금", "D10", "BS", "", ""],
        ["재무상태표", "유동부채", "예수금", "D11", "BS", "", ""],
        
        ["재무상태표", "비유동부채", "장기차입금", "D15", "BS", "", ""],
        ["재무상태표", "비유동부채", "퇴직급여충당부채", "D16", "BS", "", ""],
        ["재무상태표", "비유동부채", "기타비유동부채", "D17", "BS", "", ""],
        
        ["재무상태표", "자본", "자본금", "D21", "BS", "", ""],
        ["재무상태표", "자본", "자본잉여금", "D22", "BS", "", ""],
        ["재무상태표", "자본", "이익잉여금", "D23", "BS", "", ""],
        ["재무상태표", "자본", "기타포괄손익누계액", "D24", "BS", "", ""]
    ]
    
    # DataFrame 생성
    columns = ["파일명", "섹션", "계정명", "셀주소", "데이터소스", "제목셀", "제목템플릿"]
    df = pd.DataFrame(mapping_data, columns=columns)
    
    # Excel 파일 생성
    wb = Workbook()
    
    # 1. 셀매핑 시트
    ws_mapping = wb.active
    ws_mapping.title = "셀매핑"
    
    # 데이터 입력
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_mapping.append(r)
    
    # 헤더 스타일 적용
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws_mapping[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 컬럼 너비 조정
    column_widths = [15, 15, 20, 10, 12, 10, 35]
    for i, width in enumerate(column_widths, 1):
        ws_mapping.column_dimensions[chr(64 + i)].width = width
    
    # 2. 사용법 시트
    ws_usage = wb.create_sheet("사용법")
    
    usage_data = [
        ["컬럼명", "설명", "예시", "필수여부"],
        ["시트명", "Excel 보고서의 시트명 (실제 템플릿과 동일해야 함)", "경영실적요약, 반반장표, 재무상태표", "필수"],
        ["섹션", "같은 시트 내에서의 데이터 구분", "당월데이터, 전년동월데이터, 판관비상세", "필수"],
        ["계정명", "계정과목매핑표의 '보고서계정명'과 정확히 일치", "매출액, 급여, 현금및현금성자산", "필수"],
        ["셀주소", "값이 입력될 Excel 셀 위치", "B5, C10, D15", "필수"],
        ["데이터소스", "IS(손익계산서) 또는 BS(재무상태표)", "IS, BS", "필수"],
        ["제목셀", "제목이 입력될 셀 위치 (선택사항)", "A1", "선택"],
        ["제목템플릿", "제목 템플릿 {year}, {month} 사용 가능", "{year}년 {month}월 경영실적요약", "선택"],
        ["", "", "", ""],
        ["사용 시 주의사항", "", "", ""],
        ["1. 시트명", "실제 Excel 템플릿의 시트명과 정확히 일치해야 합니다", "", ""],
        ["2. 계정명", "계정과목매핑표의 '보고서계정명' 컬럼과 정확히 일치해야 합니다", "", ""],
        ["3. 셀주소", "Excel 형식 (A1, B5, C10 등)으로 입력하세요", "", ""],
        ["4. 데이터소스", "손익계산서는 IS, 재무상태표는 BS로 입력하세요", "", ""],
        ["5. 제목", "제목이 필요 없으면 제목셀과 제목템플릿을 비워두세요", "", ""],
        ["", "", "", ""],
        ["수정 방법", "", "", ""],
        ["1. 셀주소 변경", "실제 템플릿에 맞게 B5 → D7 등으로 변경", "", ""],
        ["2. 계정 추가", "새 행에 계정명, 셀주소 등을 입력", "", ""],
        ["3. 섹션 추가", "새로운 섹션명으로 데이터 그룹 생성", "", ""],
        ["4. 시트 추가", "새로운 시트명으로 별도 매핑 생성", "", ""]
    ]
    
    for row in usage_data:
        ws_usage.append(row)
    
    # 사용법 시트 스타일
    for cell in ws_usage[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 사용법 시트 컬럼 너비
    usage_widths = [15, 50, 40, 10]
    for i, width in enumerate(usage_widths, 1):
        ws_usage.column_dimensions[chr(64 + i)].width = width
    
    # 3. 예시 시트 (실제 매핑 예시)
    ws_example = wb.create_sheet("매핑예시")
    
    example_data = [
        ["실제 사용 예시", "", "", ""],
        ["", "", "", ""],
        ["상황: B5 셀에 매출액이 들어가야 하는 경우", "", "", ""],
        ["시트명", "섹션", "계정명", "셀주소"],
        ["경영실적요약", "당월데이터", "매출액", "B5"],
        ["", "", "", ""],
        ["상황: 새로운 계정 '렌탈비용'을 B26에 추가하는 경우", "", "", ""],
        ["시트명", "섹션", "계정명", "셀주소"],
        ["반반장표", "판관비상세", "렌탈비용", "B26"],
        ["", "", "", ""],
        ["상황: 제목을 A1에 넣고 싶은 경우", "", "", ""],
        ["시트명", "섹션", "계정명", "셀주소", "데이터소스", "제목셀", "제목템플릿"],
        ["경영실적요약", "당월데이터", "매출액", "B5", "IS", "A1", "{year}년 {month}월 경영실적요약"],
        ["", "", "", "", "", "", ""],
        ["결과: A1에 '2024년 11월 경영실적요약' 자동 입력", "", "", ""]
    ]
    
    for row in example_data:
        ws_example.append(row)
    
    # 예시 시트 스타일
    ws_example['A1'].font = Font(bold=True, size=14)
    ws_example['A3'].font = Font(bold=True, color="0066CC")
    ws_example['A7'].font = Font(bold=True, color="0066CC")
    ws_example['A11'].font = Font(bold=True, color="0066CC")
    ws_example['A15'].font = Font(bold=True, color="FF6600")
    
    # 파일 저장
    filename = "셀매핑.xlsx"
    wb.save(filename)
    print(f"✅ {filename} 파일이 생성되었습니다!")
    print(f"📁 현재 폴더에서 파일을 확인하세요: {filename}")
    
    return filename

# 실행
if __name__ == "__main__":
    create_cell_mapping_excel()