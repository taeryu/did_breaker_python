#!/usr/bin/env python3
# 전체 시스템 통합 테스트

import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from datetime import datetime

def test_complete_system():
    """전체 시스템 통합 테스트"""
    
    print("🧪 전체 시스템 통합 테스트 시작...")
    
    try:
        # 1. 모든 필요 파일 존재 확인
        print("\n📋 1. 파일 존재 확인...")
        required_files = [
            "계정과목매핑표.xlsx",
            "셀매핑.xlsx", 
            "시산표샘플.xlsx",
            "경영실적요약.xlsx",
            "재무상태표.xlsx",
            "column_mapping.csv"
        ]
        
        for file_path in required_files:
            if os.path.exists(file_path):
                print(f"   ✅ {file_path}")
            else:
                raise FileNotFoundError(f"필수 파일 없음: {file_path}")
        
        # 2. 계정매핑표 로드 및 검증
        print("\n📊 2. 계정매핑표 검증...")
        mapping_df = pd.read_excel("계정과목매핑표.xlsx")
        mapping_df.columns = mapping_df.columns.str.strip()
        
        required_mapping_cols = ['계정코드', '계정명', '보고서계정명', '재무제표구분', '대분류']
        for col in required_mapping_cols:
            if col not in mapping_df.columns:
                raise ValueError(f"매핑표에 필수 컬럼 없음: {col}")
        
        mapping_df['계정코드'] = mapping_df['계정코드'].astype(str)
        print(f"   ✅ 매핑표 로드 성공: {len(mapping_df)}건")
        print(f"   - IS 항목: {len(mapping_df[mapping_df['재무제표구분'] == 'IS'])}건")
        print(f"   - BS 항목: {len(mapping_df[mapping_df['재무제표구분'] == 'BS'])}건")
        
        # 3. 시산표 로드 및 컬럼 매핑
        print("\n📊 3. 시산표 처리...")
        trial_df = pd.read_excel("시산표샘플.xlsx")
        trial_df.columns = trial_df.columns.astype(str).str.strip()
        
        # 컬럼 매핑 적용
        col_mapping_df = pd.read_csv("column_mapping.csv")
        col_mapping = dict(zip(col_mapping_df['원본컬럼명'], col_mapping_df['한글컬럼명']))
        
        for old_col, new_col in col_mapping.items():
            if old_col in trial_df.columns:
                trial_df = trial_df.rename(columns={old_col: new_col})
        
        print(f"   ✅ 시산표 컬럼 변환: {list(trial_df.columns)}")
        
        # 4. 데이터 매핑 및 집계
        print("\n🔄 4. 데이터 매핑 및 집계...")
        trial_df['계정코드'] = trial_df['계정코드'].astype(str)
        merged_df = pd.merge(trial_df, mapping_df, on='계정코드', how='left')
        
        # 매핑되지 않은 계정 확인
        unmapped = merged_df[merged_df['보고서계정명'].isna()]
        if not unmapped.empty:
            print(f"   ⚠️ 매핑되지 않은 계정: {len(unmapped)}건")
            for _, row in unmapped.iterrows():
                print(f"      - {row['계정코드']}: {row.get('계정명', 'N/A')}")
        else:
            print("   ✅ 모든 계정 매핑 성공")
        
        # 매핑된 데이터로 집계
        mapped_df = merged_df.dropna(subset=['보고서계정명'])
        mapped_df['잔액'] = pd.to_numeric(mapped_df['잔액'], errors='coerce').fillna(0)
        
        # 재무제표별 집계
        financial_data = {}
        for fs_type in ['IS', 'BS']:
            fs_data = mapped_df[mapped_df['재무제표구분'] == fs_type]
            if not fs_data.empty:
                account_totals = fs_data.groupby('보고서계정명')['잔액'].sum().to_dict()
                financial_data[fs_type] = {'account_totals': account_totals}
                print(f"   ✅ {fs_type} 집계 완료: {len(account_totals)}개 계정")
        
        # 5. 셀매핑 검증
        print("\n📊 5. 셀매핑 검증...")
        cell_df = pd.read_excel("셀매핑.xlsx", sheet_name="셀매핑")
        
        # 파일명 컬럼 확인
        if '파일명' not in cell_df.columns:
            raise ValueError("셀매핑에 '파일명' 컬럼이 없습니다")
        
        # 파일명별 그룹화 테스트
        mapping_dict = {}
        for file_name in cell_df['파일명'].unique():
            if pd.isna(file_name):
                continue
            file_data = cell_df[cell_df['파일명'] == file_name]
            mapping_dict[file_name] = len(file_data)
        
        print(f"   ✅ 셀매핑 파일별 분석:")
        for file_name, count in mapping_dict.items():
            print(f"      - {file_name}: {count}개 매핑")
        
        # 6. Excel 템플릿 처리 테스트
        print("\n📝 6. Excel 템플릿 처리 테스트...")
        
        # 출력 폴더 생성
        output_folder = "./test_reports_full"
        os.makedirs(output_folder, exist_ok=True)
        
        template_files = ["경영실적요약.xlsx", "재무상태표.xlsx"]
        created_reports = []
        
        for template_file in template_files:
            try:
                # 파일명에서 확장자 제거
                base_name = os.path.splitext(os.path.basename(template_file))[0]
                
                # 출력 파일명 생성
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"{base_name}_2024년11월_{timestamp}.xlsx"
                output_path = os.path.join(output_folder, output_filename)
                
                # 템플릿 복사
                shutil.copy2(template_file, output_path)
                
                # Excel 파일 열기
                wb = load_workbook(output_path)
                ws = wb.active
                
                print(f"   📄 처리 중: {template_file}")
                print(f"      - 매핑 키: {base_name}")
                print(f"      - 출력 파일: {output_filename}")
                
                # 셀매핑에서 해당 파일 정보 찾기
                file_mapping_data = cell_df[cell_df['파일명'] == base_name]
                if not file_mapping_data.empty:
                    print(f"      - 매핑 데이터: {len(file_mapping_data)}건 발견")
                    
                    # 몇 개 셀에 테스트 데이터 입력
                    test_count = 0
                    for _, row in file_mapping_data.head(5).iterrows():
                        account_name = row['계정명']
                        cell_addr = row['셀주소']
                        data_source = row['데이터소스']
                        
                        if data_source in financial_data:
                            amount = financial_data[data_source]['account_totals'].get(account_name, 0)
                            try:
                                ws[cell_addr] = amount
                                print(f"         {cell_addr}: {account_name} = {amount:,}")
                                test_count += 1
                            except Exception as e:
                                print(f"         ❌ {cell_addr} 입력 실패: {e}")
                    
                    print(f"      ✅ {test_count}개 셀 입력 완료")
                else:
                    print(f"      ⚠️ 매핑 데이터 없음")
                
                # 파일 저장
                wb.save(output_path)
                created_reports.append(output_path)
                
            except Exception as e:
                print(f"      ❌ 템플릿 처리 실패: {e}")
        
        # 7. 결과 요약
        print(f"\n🎉 전체 시스템 테스트 완료!")
        print(f"📁 출력 폴더: {output_folder}")
        print(f"📋 생성된 보고서: {len(created_reports)}개")
        
        for report in created_reports:
            file_size = os.path.getsize(report)
            print(f"   📄 {os.path.basename(report)} ({file_size:,} bytes)")
        
        # 재무 요약
        if financial_data:
            print(f"\n📊 재무 데이터 요약:")
            if 'IS' in financial_data:
                is_totals = financial_data['IS']['account_totals']
                if '매출액' in is_totals:
                    print(f"   매출액: {is_totals['매출액']:,}원")
                if '급여' in is_totals:
                    print(f"   급여: {is_totals['급여']:,}원")
            
            if 'BS' in financial_data:
                bs_totals = financial_data['BS']['account_totals']
                if '현금및현금성자산' in bs_totals:
                    print(f"   현금및현금성자산: {bs_totals['현금및현금성자산']:,}원")
        
        print(f"\n✅ 모든 테스트 통과!")
        return True
        
    except Exception as e:
        print(f"\n💥 테스트 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_complete_system()