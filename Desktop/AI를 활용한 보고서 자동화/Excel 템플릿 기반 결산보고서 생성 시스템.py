# 간소화된 Excel 템플릿 기반 보고서 생성 시스템
# main_simple.py

import pandas as pd
from datetime import datetime
import logging
import os
from openpyxl import load_workbook
import shutil
import tkinter as tk
from tkinter import filedialog

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def select_file(title, filetypes=[("Excel files", "*.xlsx *.xls")]):
    """파일 선택 팝업"""
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    root.lift()
    root.attributes('-topmost', True)
    
    filename = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    root.destroy()
    return filename

def select_multiple_files(title, filetypes=[("Excel files", "*.xlsx *.xls")]):
    """여러 파일 선택 팝업"""
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    root.lift()
    root.attributes('-topmost', True)
    
    filenames = filedialog.askopenfilenames(
        title=title,
        filetypes=filetypes
    )
    root.destroy()
    return list(filenames)

def select_folder(title="폴더 선택"):
    """폴더 선택 팝업"""
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    root.lift()
    root.attributes('-topmost', True)
    
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return folder

class AccountMappingManager:
    """기존 Excel 매핑테이블 관리"""
    
    def __init__(self, mapping_file_path=None):
        if mapping_file_path is None:
            print("계정과목매핑표 Excel 파일을 선택하세요...")
            mapping_file_path = select_file("계정과목매핑표 Excel 파일 선택")
            if not mapping_file_path:
                raise Exception("계정과목매핑표 파일이 선택되지 않았습니다.")
        
        self.mapping_file_path = mapping_file_path
        self.mapping_df = None
        self.load_mapping()
    
    def load_mapping(self):
        """Excel 매핑테이블 로드"""
        try:
            self.mapping_df = pd.read_excel(self.mapping_file_path)
            self.mapping_df.columns = self.mapping_df.columns.str.strip()
            
            # 계정코드를 문자열로 변환 (원본 그대로 유지)
            if '계정코드' in self.mapping_df.columns:
                self.mapping_df['계정코드'] = self.mapping_df['계정코드'].astype(str)
            
            logging.info(f"계정매핑표 로드 완료: {len(self.mapping_df)}건")
            print(f"✅ 계정매핑표 로드 완료: {len(self.mapping_df)}건")
            
        except Exception as e:
            logging.error(f"매핑테이블 로드 실패: {e}")
            raise
    
    def get_mapping_df(self):
        return self.mapping_df

class SAPDataProcessor:
    """SAP 데이터 처리 및 매핑 적용"""
    
    def __init__(self, mapping_manager):
        self.mapping_manager = mapping_manager
        self.mapping_df = mapping_manager.get_mapping_df()
    
    def load_column_mapping(self):
        """CSV 파일에서 컬럼 매핑 로드"""
        try:
            mapping_file = os.path.join(os.path.dirname(__file__), 'column_mapping.csv')
            if os.path.exists(mapping_file):
                df = pd.read_csv(mapping_file)
                return dict(zip(df['원본컬럼명'], df['한글컬럼명']))
            else:
                # 기본 매핑 반환
                return {
                    'G/L Account': '계정코드',
                    'Account': '계정코드', 
                    'GL Account': '계정코드',
                    'Account Name': '계정명',
                    'G/L Account Name': '계정명',
                    'Debit': '차변',
                    'Credit': '대변', 
                    'Balance': '잔액',
                    'Amount': '금액'
                }
        except Exception as e:
            logging.warning(f"컬럼 매핑 로드 실패, 기본값 사용: {e}")
            return {
                'G/L Account': '계정코드',
                'Account': '계정코드', 
                'GL Account': '계정코드',
                'Account Name': '계정명',
                'G/L Account Name': '계정명',
                'Debit': '차변',
                'Credit': '대변', 
                'Balance': '잔액',
                'Amount': '금액'
            }

    def load_trial_balance(self, file_path):
        """시산표 파일 로드 - 강화된 버전"""
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"시산표 파일을 찾을 수 없습니다: {file_path}")
            
            # Excel 파일 읽기
            df = pd.read_excel(file_path, skiprows=0)
            
            # 빈 데이터프레임 확인
            if df.empty:
                raise ValueError("시산표 데이터가 비어있습니다")
            
            # 컬럼명 정리
            df.columns = df.columns.astype(str).str.strip()
            
            # CSV에서 컬럼 매핑 로드
            column_mapping = self.load_column_mapping()
            
            for old_col, new_col in column_mapping.items():
                if old_col in df.columns:
                    df = df.rename(columns={old_col: new_col})
            
            # 필수 컬럼 확인
            if '계정코드' not in df.columns:
                raise ValueError("계정코드 컬럼을 찾을 수 없습니다")
            
            # 데이터 정제
            df = df.dropna(subset=['계정코드'])
            df['계정코드'] = df['계정코드'].astype(str)
            
            # 금액 컬럼 처리
            amount_columns = ['차변', '대변', '잔액', '금액']
            for col in amount_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # 잔액 계산
            if '잔액' not in df.columns and '차변' in df.columns and '대변' in df.columns:
                df['잔액'] = df['차변'] - df['대변']
            
            logging.info(f"시산표 데이터 로드 완료: {len(df)}건")
            return df
            
        except Exception as e:
            logging.error(f"시산표 로드 실패: {e}")
            return None
    
    def calculate_financial_data(self, trial_balance_df):
        """매핑 테이블을 이용해 재무데이터 계산"""
        
        # pandas merge로 vlookup 방식 매핑
        merged_df = pd.merge(trial_balance_df, self.mapping_df, on='계정코드', how='left')
        
        # 매핑되지 않은 계정 확인
        unmapped = merged_df[merged_df['보고서계정명'].isna()]
        if not unmapped.empty:
            logging.warning(f"매핑되지 않은 계정 {len(unmapped)}건 존재")
            print(f"⚠️ 매핑되지 않은 계정 {len(unmapped)}건:")
            for _, row in unmapped.iterrows():
                print(f"   계정코드: {row['계정코드']}, 계정명: {row.get('계정명', 'N/A')}")
        
        # 매핑된 데이터만 처리
        mapped_df = merged_df.dropna(subset=['보고서계정명'])
        
        # 계정별 집계
        financial_data = {}
        
        # 보고서계정명별 합계 계산
        account_totals = mapped_df.groupby('보고서계정명')['잔액'].sum().to_dict()
        
        # 재무제표 구분별 계산
        for fs_type in ['IS', 'BS']:
            fs_data = mapped_df[mapped_df['재무제표구분'] == fs_type]
            if not fs_data.empty:
                category_totals = fs_data.groupby('대분류')['잔액'].sum().to_dict()
                financial_data[fs_type] = {
                    'account_totals': fs_data.groupby('보고서계정명')['잔액'].sum().to_dict(),
                    'category_totals': category_totals
                }
        
        return financial_data, account_totals

class ExcelTemplateProcessor:
    """Excel 템플릿 기반 보고서 생성"""
    
    def __init__(self, cell_mapping_file_path=None):
        if cell_mapping_file_path is None:
            print("셀매핑 Excel 파일을 선택하세요...")
            cell_mapping_file_path = select_file("셀매핑 Excel 파일 선택")
            if not cell_mapping_file_path:
                raise Exception("셀매핑 파일이 선택되지 않았습니다.")
        
        self.cell_mapping_file = cell_mapping_file_path
        self.cell_mapping = self.load_cell_mapping()
    
    def load_cell_mapping(self):
        """셀 매핑 정보를 Excel 파일에서 로드"""
        try:
            if not os.path.exists(self.cell_mapping_file):
                raise FileNotFoundError(f"셀매핑 파일을 찾을 수 없습니다: {self.cell_mapping_file}")
            
            df = pd.read_excel(self.cell_mapping_file, sheet_name='셀매핑')
            cell_mapping = self.convert_excel_to_mapping_dict(df)
            logging.info(f"셀 매핑 정보 로드 완료: {self.cell_mapping_file}")
            print(f"✅ 셀매핑 정보 로드 완료: {len(cell_mapping)}개 파일")
            return cell_mapping
                    
        except Exception as e:
            logging.error(f"셀 매핑 로드 실패: {e}")
            raise
    
    def convert_excel_to_mapping_dict(self, df):
        """Excel DataFrame을 매핑 딕셔너리로 변환"""
        mapping_dict = {}
        
        # 파일명별로 그룹화
        for file_name in df['파일명'].unique():
            if pd.isna(file_name):
                continue
                
            file_data = df[df['파일명'] == file_name]
            mapping_dict[file_name] = {}
            
            # _config 정보 설정 (첫 번째 행에서 가져오기)
            first_row = file_data.iloc[0]
            config = {}
            if pd.notna(first_row.get('데이터소스')):
                config['data_source'] = first_row['데이터소스']
            if pd.notna(first_row.get('제목셀')):
                config['title_cell'] = first_row['제목셀']
            if pd.notna(first_row.get('제목템플릿')):
                config['title_template'] = first_row['제목템플릿']
            
            if config:
                mapping_dict[file_name]['_config'] = config
            
            # 섹션별로 그룹화
            for section_name in file_data['섹션'].unique():
                if pd.isna(section_name):
                    continue
                    
                section_data = file_data[file_data['섹션'] == section_name]
                mapping_dict[file_name][section_name] = {}
                
                # 계정명과 셀주소 매핑
                for _, row in section_data.iterrows():
                    if pd.notna(row['계정명']) and pd.notna(row['셀주소']):
                        mapping_dict[file_name][section_name][row['계정명']] = row['셀주소']
        
        return mapping_dict
    
    def create_reports_from_templates(self, template_files, financial_data, previous_data, year, month, output_folder):
        """여러 템플릿에서 보고서 생성 - 첫 번째 시트에만 데이터 입력"""
        
        created_reports = []
        
        for template_file in template_files:
            try:
                # 파일명에서 확장자 제거
                base_name = os.path.splitext(os.path.basename(template_file))[0]
                
                # 출력 파일명 생성
                output_filename = f"{base_name}_{year}년{month:02d}월_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_path = os.path.join(output_folder, output_filename)
                
                # 템플릿 복사
                shutil.copy2(template_file, output_path)
                
                # 워크북 열기
                wb = load_workbook(output_path)
                
                # 첫 번째 시트 가져오기
                first_sheet_name = wb.sheetnames[0]
                
                # 매핑 정보가 있는 경우에만 데이터 입력
                file_mapping_key = base_name
                if file_mapping_key in self.cell_mapping:
                    success = self.fill_sheet_data(wb, first_sheet_name, self.cell_mapping[file_mapping_key], 
                                                 financial_data, previous_data, year, month)
                    if success:
                        wb.save(output_path)
                        created_reports.append(output_path)
                        logging.info(f"보고서 생성 완료: {output_path}")
                        print(f"✅ 보고서 생성: {os.path.basename(output_path)}")
                    else:
                        logging.warning(f"보고서 데이터 입력 실패: {output_path}")
                        print(f"⚠️ 데이터 입력 실패: {os.path.basename(output_path)}")
                else:
                    logging.warning(f"매핑 정보를 찾을 수 없음: {file_mapping_key}")
                    print(f"⚠️ 매핑 정보 없음: {file_mapping_key}")
                    # 매핑 정보가 없어도 템플릿은 복사해둠
                    wb.save(output_path)
                    created_reports.append(output_path)
                
            except Exception as e:
                logging.error(f"보고서 생성 실패 ({template_file}): {e}")
                print(f"❌ 보고서 생성 실패: {os.path.basename(template_file)} - {e}")
                continue
        
        return created_reports
    
    def fill_sheet_data(self, workbook, sheet_name, sheet_mapping, financial_data, previous_data=None, year=None, month=None):
        """시트에 데이터 입력 - 첫 번째 시트 전용"""
        
        try:
            # 시트 가져오기
            if sheet_name not in workbook.sheetnames:
                logging.warning(f"시트를 찾을 수 없습니다: {sheet_name}")
                return False
            
            ws = workbook[sheet_name]
            
            # 설정 정보 추출
            config = sheet_mapping.get("_config", {})
            data_source = config.get("data_source", "IS")  # IS 또는 BS
            title_cell = config.get("title_cell")
            title_template = config.get("title_template")
            
            # 제목 업데이트 (설정되어 있는 경우)
            if title_cell and title_template and year and month:
                try:
                    title = title_template.format(year=year, month=month)
                    ws[title_cell] = title
                    logging.debug(f"제목 업데이트: {title} → {title_cell}")
                except Exception as e:
                    logging.warning(f"제목 업데이트 실패: {e}")
            
            # 데이터 소스에서 계정별 금액 가져오기
            current_data = financial_data.get(data_source, {}).get('account_totals', {})
            previous_data_dict = {}
            if previous_data:
                previous_data_dict = previous_data.get(data_source, {}).get('account_totals', {})
            
            # 각 섹션별로 데이터 입력
            for section_name, section_mapping in sheet_mapping.items():
                if section_name.startswith("_"):  # _config 같은 설정 섹션은 스킵
                    continue
                
                # 섹션별 데이터 선택
                if section_name == "전년동월데이터" and previous_data_dict:
                    # 전년 데이터 사용
                    data_to_use = previous_data_dict
                    logging.debug(f"전년 데이터 입력 섹션: {section_name}")
                else:
                    # 당월 데이터 사용 (기본)
                    data_to_use = current_data
                    logging.debug(f"당월 데이터 입력 섹션: {section_name}")
                
                # 해당 섹션의 계정별 데이터 입력
                for account_name, cell_address in section_mapping.items():
                    amount = data_to_use.get(account_name, 0)
                    try:
                        ws[cell_address] = amount
                        logging.debug(f"데이터 입력: {section_name}.{account_name} = {amount:,} → {cell_address}")
                    except Exception as e:
                        logging.warning(f"셀 입력 실패: {account_name} → {cell_address}, 오류: {e}")
            
            logging.info(f"시트 데이터 입력 완료: {sheet_name}")
            return True
            
        except Exception as e:
            logging.error(f"시트 데이터 입력 실패: {e}")
            return False

class MonthlyClosingProcessor:
    """월마감 메인 처리 클래스"""
    
    def __init__(self, mapping_file_path=None, cell_mapping_file_path=None):
        self.mapping_manager = AccountMappingManager(mapping_file_path)
        self.data_processor = SAPDataProcessor(self.mapping_manager)
        self.template_processor = ExcelTemplateProcessor(cell_mapping_file_path)
    
    def process_monthly_closing(self, sap_file_path, template_files, year, month, 
                               output_folder="./reports", previous_file_path=None):
        """월마감 보고서 생성 메인 프로세스"""
        
        try:
            print(f"\n🔄 {year}년 {month}월 월마감 처리 시작...")
            
            # 1. 시산표 데이터 로드 및 계산
            print("📊 재무데이터 계산 중...")
            trial_balance_df = self.data_processor.load_trial_balance(sap_file_path)
            if trial_balance_df is None:
                raise Exception("시산표 데이터 로드 실패")
            
            current_financial_data, account_totals = self.data_processor.calculate_financial_data(trial_balance_df)
            print(f"✅ 당월 데이터 처리 완료: {len(account_totals)}개 계정")
            
            # 2. 전년 동월 데이터 로드 (파일이 지정된 경우)
            previous_financial_data = None
            if previous_file_path and os.path.exists(previous_file_path):
                print("📊 전년 동월 데이터 처리 중...")
                prev_trial_balance = self.data_processor.load_trial_balance(previous_file_path)
                if prev_trial_balance is not None:
                    previous_financial_data, _ = self.data_processor.calculate_financial_data(prev_trial_balance)
                    print(f"✅ 전년 동월 데이터 로드 완료")
            
            # 3. 출력 폴더 생성
            os.makedirs(output_folder, exist_ok=True)
            
            # 4. 템플릿 기반 보고서 생성
            print("📋 보고서 생성 중...")
            created_reports = self.template_processor.create_reports_from_templates(
                template_files, current_financial_data, previous_financial_data, year, month, output_folder
            )
            
            if not created_reports:
                raise Exception("보고서 생성 실패")
            
            return {
                'success': True,
                'created_reports': created_reports,
                'summary': self._generate_summary(current_financial_data)
            }
            
        except Exception as e:
            logging.error(f"월마감 처리 실패: {e}")
            return {'success': False, 'error': str(e)}
    
    def _generate_summary(self, financial_data):
        """처리 결과 요약"""
        summary = {}
        
        if 'IS' in financial_data:
            is_totals = financial_data['IS']['account_totals']
            summary['revenue'] = is_totals.get('매출액', 0)
            summary['operating_income'] = is_totals.get('영업이익', 0)
        
        if 'BS' in financial_data:
            bs_totals = financial_data['BS']['account_totals']
            summary['total_assets'] = bs_totals.get('자산총계', 0)
        
        return summary

def main():
    """메인 실행 함수"""
    print("🏢 SAP 월마감 자동화 시스템")
    print("=" * 50)
    
    try:
        # 1. 프로세서 초기화 (여기서 파일 선택 팝업 뜸)
        print("\n1️⃣ 설정 파일 선택...")
        processor = MonthlyClosingProcessor()
        
        # 2. SAP 시산표 파일 선택
        print("\n2️⃣ SAP 시산표 파일 선택...")
        sap_file = input("SAP 시산표 파일 경로를 입력하거나 Enter를 눌러 파일 선택: ").strip()
        if not sap_file:
            sap_file = select_file("SAP 시산표 파일 선택")
        
        if not sap_file or not os.path.exists(sap_file):
            print("❌ SAP 시산표 파일이 선택되지 않았습니다.")
            return
        
        # 3. 전년 데이터 파일은 사용하지 않음
        print("\n3️⃣ 전년 동월 데이터는 기존 숫자를 그대로 사용합니다.")
        previous_file = None
        
        # 4. 년월 입력
        print("\n4️⃣ 처리 년월 입력...")
        try:
            year = int(input(f"년도 (기본: {datetime.now().year}): ") or datetime.now().year)
            month = int(input(f"월 (기본: {datetime.now().month}): ") or datetime.now().month)
        except ValueError:
            print("❌ 올바른 숫자를 입력하세요.")
            return
        
        # 5. 템플릿 파일들 선택
        print("\n5️⃣ Excel 템플릿 파일들 선택...")
        template_files = select_multiple_files("Excel 템플릿 파일들 선택 (여러 개 선택 가능)")
        if not template_files:
            print("❌ 템플릿 파일이 선택되지 않았습니다.")
            return
        
        print(f"선택된 템플릿: {len(template_files)}개")
        for i, template in enumerate(template_files, 1):
            print(f"  {i}. {os.path.basename(template)}")
        
        # 6. 출력 폴더 선택
        print("\n6️⃣ 출력 폴더 선택...")
        output_folder = input("출력 폴더 경로 (기본: ./reports): ").strip() or "./reports"
        
        # 7. 처리 실행
        print("\n🚀 처리 시작...")
        result = processor.process_monthly_closing(
            sap_file, template_files, year, month, output_folder, previous_file
        )
        
        # 8. 결과 출력
        if result['success']:
            print("\n🎉 월마감 보고서 생성 완료!")
            print(f"📁 출력 폴더: {output_folder}")
            print(f"📋 생성된 보고서: {len(result['created_reports'])}개")
            
            for report in result['created_reports']:
                print(f"   📄 {os.path.basename(report)}")
            
            # 요약 정보
            summary = result.get('summary', {})
            if summary:
                print("\n📊 재무 요약:")
                if 'revenue' in summary:
                    print(f"   매출액: {summary['revenue']:,}원")
                if 'operating_income' in summary:
                    print(f"   영업이익: {summary['operating_income']:,}원")
                if 'total_assets' in summary:
                    print(f"   총자산: {summary['total_assets']:,}원")
        else:
            print(f"\n❌ 처리 실패: {result['error']}")
    
    except Exception as e:
        print(f"\n❌ 예외 발생: {str(e)}")

if __name__ == "__main__":
    main()